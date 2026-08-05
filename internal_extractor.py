import hashlib
import random
import time
from datetime import datetime
from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from sitemap_checker import collect_rendered_links_browser, suggest_anchor_text


def internal_extractor_output_filename() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"Extracted_Internal_Pages_{timestamp}.xlsx"


def extract_internal_links_tool() -> None:
    st.title("Internal Link & Anchor Extractor")
    st.caption(
        "Extract all internal links and their exact anchor text from one or more pages using a headless browser."
    )

    urls_text = st.text_area(
        "Source URLs (one per line)",
        placeholder="https://www.example.com/page1/\nhttps://www.example.com/page2/",
    )

    signature = None
    if urls_text:
        signature = hashlib.sha256(urls_text.strip().encode("utf-8")).hexdigest()

        if st.session_state.get("extractor_input_signature") != signature:
            for key in ("extractor_output", "extractor_filename", "extractor_stats"):
                st.session_state.pop(key, None)
            st.session_state["extractor_input_signature"] = signature

    if urls_text and st.button("Extract Links", type="primary", use_container_width=True):
        urls = [u.strip() for u in urls_text.splitlines() if u.strip()]
        if not urls:
            st.error("Please enter at least one URL.")
        else:
            progress_bar = st.progress(0, text="Starting extraction…")

            all_extracted_links = []
            source_pages = []

            try:
                for i, url in enumerate(urls):
                    progress_bar.progress(
                        i / len(urls), text=f"Processing page {i + 1} of {len(urls)}…"
                    )

                    # Fetch links using headless browser
                    links = collect_rendered_links_browser(url)

                    # Derive page name
                    page_name = suggest_anchor_text(url)

                    source_pages.append({"url": url, "page_name": page_name})

                    for link in links:
                        all_extracted_links.append(
                            {
                                "source_page_name": page_name,
                                "url": link["url"],
                                "anchor": link["anchor"],
                            }
                        )

                    if i < len(urls) - 1:
                        time.sleep(random.uniform(2, 5))

                progress_bar.progress(1.0, text="Extraction complete")

                # Now build the excel file
                template_path = Path(__file__).with_name("Extracted_Internal_Pages.xlsx")
                if not template_path.exists():
                    raise ValueError(
                        "Template Extracted_Internal_Pages.xlsx is missing from the deployed app."
                    )

                wb = load_workbook(template_path)

                if "Internal Links" not in wb.sheetnames or "Source URL" not in wb.sheetnames:
                    raise ValueError(
                        "Template does not contain 'Internal Links' and 'Source URL' sheets."
                    )

                internal_links_ws = wb["Internal Links"]
                source_url_ws = wb["Source URL"]

                for idx, link_data in enumerate(all_extracted_links, start=2):
                    internal_links_ws.cell(row=idx, column=1).value = link_data["url"]
                    internal_links_ws.cell(row=idx, column=2).value = link_data["anchor"]
                    internal_links_ws.cell(row=idx, column=3).value = link_data["source_page_name"]

                for idx, source_data in enumerate(source_pages, start=2):
                    source_url_ws.cell(row=idx, column=1).value = source_data["url"]
                    source_url_ws.cell(row=idx, column=2).value = source_data["page_name"]

                output = BytesIO()
                wb.save(output)
                output_bytes = output.getvalue()

                st.session_state["extractor_output"] = output_bytes
                st.session_state["extractor_filename"] = internal_extractor_output_filename()
                st.session_state["extractor_stats"] = {
                    "source_pages": len(source_pages),
                    "internal_links": len(all_extracted_links),
                }

            except Exception as exc:
                progress_bar.empty()
                st.error(str(exc))

    if st.session_state.get("extractor_stats"):
        stats = st.session_state["extractor_stats"]
        col1, col2 = st.columns(2)
        col1.metric("Total source pages processed", stats["source_pages"])
        col2.metric("Total internal links found", stats["internal_links"])

        st.download_button(
            "Download Extracted Links (.xlsx)",
            data=st.session_state["extractor_output"],
            file_name=st.session_state["extractor_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
