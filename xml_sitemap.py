import gzip
import xml.etree.ElementTree as ET
from io import BytesIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import urlsplit
from urllib.request import Request, urlopen

from openpyxl import load_workbook


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/150.0.0.0 Safari/537.36"
)

def _fetch_recursive(url: str, max_depth: int, current_depth: int, seen_sitemaps: set[str], out_urls: list[str]) -> None:
    if current_depth > max_depth:
        return
    if url in seen_sitemaps:
        return
    seen_sitemaps.add(url)
    
    request = Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Cache-Control": "no-cache",
        },
    )
    
    try:
        with urlopen(request, timeout=30) as response:
            data = response.read(25_000_000)
    except HTTPError as exc:
        raise ValueError(f"Sitemap URL returned HTTP {exc.code}: {url}") from exc
    except (URLError, TimeoutError, OSError) as exc:
        raise ValueError(f"Sitemap URL could not be fetched: {url}. Error: {exc}") from exc

    # Check if gzip
    if url.casefold().endswith(".gz") or data[:2] == b'\x1f\x8b':
        try:
            data = gzip.decompress(data)
        except Exception as exc:
            raise ValueError(f"Failed to decompress gzip sitemap: {url}") from exc

    try:
        root = ET.fromstring(data)
    except ET.ParseError as exc:
        raise ValueError(f"Response from {url} is not valid XML.") from exc
        
    for elem in root.iter():
        tag = elem.tag.split("}", 1)[-1] if "}" in elem.tag else elem.tag
        if tag == "sitemap":
            for child in elem:
                child_tag = child.tag.split("}", 1)[-1] if "}" in child.tag else child.tag
                if child_tag == "loc" and child.text:
                    loc_url = child.text.strip()
                    _fetch_recursive(loc_url, max_depth, current_depth + 1, seen_sitemaps, out_urls)
                    break
        elif tag == "url":
            for child in elem:
                child_tag = child.tag.split("}", 1)[-1] if "}" in child.tag else child.tag
                if child_tag == "loc" and child.text:
                    out_urls.append(child.text.strip())
                    break

def fetch_xml_sitemap_urls(sitemap_url: str, max_depth: int = 2) -> list[str]:
    """Fetches and parses a sitemap or sitemapindex, returning a list of URLs."""
    if not sitemap_url.casefold().startswith(("http://", "https://")):
        raise ValueError("Enter a complete XML sitemap URL beginning with http:// or https://.")
    
    urls_found = []
    _fetch_recursive(sitemap_url, max_depth, 0, set(), urls_found)
    
    seen = set()
    deduped = []
    for u in urls_found:
        if u not in seen:
            seen.add(u)
            deduped.append(u)
            
    if not deduped:
        raise ValueError("The sitemap appears to be empty or contains no valid <loc> tags.")
        
    return deduped

def filter_sitemap_urls(urls: list[str], path_filter: str, country_filter: str) -> list[str]:
    """Apply AND filter based on given path and country substrings."""
    filtered = []
    p_filter = path_filter.strip().casefold() if path_filter else ""
    c_filter = country_filter.strip().casefold() if country_filter else ""
    
    for url in urls:
        url_lower = url.casefold()
        path_matches = p_filter in url_lower if p_filter else True
        country_matches = c_filter in url_lower if c_filter else True
        
        if path_matches and country_matches:
            filtered.append(url)
            
    return filtered

def generate_xml_export_workbook(urls: list[str]) -> bytes:
    template_path = Path(__file__).with_name("XML_Sitemap_Export.xlsx")
    if not template_path.exists():
        raise ValueError("XML sitemap Excel template (XML_Sitemap_Export.xlsx) is missing from the deployed app.")
        
    wb = load_workbook(template_path)
    pages_ws = wb["Pages"]
    
    # Clear existing data from A2 downwards
    for row in pages_ws.iter_rows(min_row=2, max_col=1, max_row=pages_ws.max_row):
        for cell in row:
            cell.value = None

    for idx, url in enumerate(urls, start=2):
        pages_ws.cell(row=idx, column=1).value = url
        
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
    
def xml_sitemap_output_filename(sitemap_url: str) -> str:
    parts = urlsplit(sitemap_url)
    host = parts.hostname or "sitemap"
    return f"xml_sitemap_export_{host}.xlsx"
