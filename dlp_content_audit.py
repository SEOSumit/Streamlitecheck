import io
import datetime
import concurrent.futures
import queue
import time
import openpyxl
from openpyxl.styles import PatternFill, Font
import streamlit as st
from playwright.sync_api import sync_playwright, Error as PlaywrightError

JS_EXTRACT_SCRIPT = r"""
() => {
    function isVisible(el) {
        if (!el) return false;
        const style = window.getComputedStyle(el);
        const rect = el.getBoundingClientRect();
        return style.display !== 'none' && 
               style.visibility !== 'hidden' && 
               style.opacity !== '0' && 
               rect.width > 0 && 
               rect.height > 0 && 
               el.getAttribute('aria-hidden') !== 'true' &&
               el.offsetParent !== null;
    }

    const data = {
        page_heading: "",
        product_region_found: false,
        reported_product_count: 0,
        rendered_product_cards: 0,
        sample_products: [],
        product_loading_skeleton: false,
        product_detection_method: "",
        bopc_available: false,
        bopc_char_count: 0,
        bopc_heading_count: 0,
        bopc_headings: [],
        bopc_detection_method: "",
        qa_available: false,
        qa_count: 0,
        sample_qa: [],
        qa_detection_method: "",
        component_sequence: [],
        audit_notes: [],
        debug_candidates: []
    };

    // 1. Page Heading
    const h1s = Array.from(document.querySelectorAll('h1')).filter(isVisible);
    if (h1s.length > 0) {
        data.page_heading = h1s[0].innerText.trim();
    } else if (document.title) {
        data.page_heading = document.title;
    } else {
        const metaTitle = document.querySelector('meta[name="title"]');
        if (metaTitle) data.page_heading = metaTitle.content;
    }

    // Identify Main Region
    let mainRegion = document.querySelector('main, [role="main"], #main-content');
    if (!mainRegion) mainRegion = document.body;

    // Footer Boundary
    const footerEls = Array.from(document.querySelectorAll('footer, [role="contentinfo"], [componentname="commonFooter" i], [componentName="commonFooter" i], .commonFooter'));
    let footerTop = document.body.scrollHeight;
    let footerMethod = "none";
    if (footerEls.length > 0) {
        const footer = footerEls[footerEls.length - 1];
        footerTop = footer.getBoundingClientRect().top + window.scrollY;
        footerMethod = footer.tagName.toLowerCase() + (footer.getAttribute('componentname') ? '[componentname]' : '');
    }
    data.audit_notes.push(`Footer detected via: ${footerMethod} at vertical pos: ${footerTop}`);

    // 2. Product Region
    let productRoot = mainRegion;
    const lenovoDlp = mainRegion.querySelector('[componentname="ofp-2c-mobile-new-dlp" i]');
    if (lenovoDlp && isVisible(lenovoDlp)) {
        data.product_region_found = true;
        productRoot = lenovoDlp;
        data.product_detection_method = "Site-specific selector";
        
        // Lenovo often puts count in a specific element
        const countEl = lenovoDlp.querySelector('.number-count, .product-count, [data-count]');
        if (countEl) {
            const num = parseInt(countEl.innerText.replace(/[^0-9]/g, ''));
            if (!isNaN(num)) data.reported_product_count = num;
        } else {
            data.reported_product_count = 1; // Assuming exists if region exists
        }
    } else {
        // Generic product region
        const grids = Array.from(mainRegion.querySelectorAll('.product-grid, .product-list, .product_list, .productList_container, [data-product-list], ul.products'));
        if (grids.length > 0) {
            data.product_region_found = true;
            productRoot = grids[0];
            data.product_detection_method = "Generic product grid";
            data.reported_product_count = 1;
        }
    }

    // Skeletons
    const skeletons = Array.from(mainRegion.querySelectorAll('.pc_dlp_skeleton_box, .skeleton_product_card, .skeleton_product, .skeleton'));
    data.product_loading_skeleton = skeletons.some(isVisible);

    // Rendered cards hierarchy
    let cards = [];
    if (productRoot !== document.body) {
        // 1. Known selectors
        cards = Array.from(productRoot.querySelectorAll('.product-item, .product-card, .product_item, .product_card, [data-product-id], .list-item, [data-testid="product-card"]')).filter(isVisible).filter(c => !c.className.includes('skeleton'));
        
        // 2. Structural discovery if 0
        if (cards.length === 0) {
            const links = Array.from(productRoot.querySelectorAll('a[href]')).filter(isVisible);
            const candidates = [];
            links.forEach(a => {
                let p = a.parentElement;
                while (p && p !== productRoot && p.tagName !== 'BODY') {
                    if (p.querySelector('img') && p.innerText.match(/[0-9]/) && !p.className.includes('skeleton')) {
                        candidates.push(p);
                        break;
                    }
                    p = p.parentElement;
                }
            });
            const classCounts = new Map();
            candidates.forEach(c => {
                const key = c.tagName + "|" + c.className;
                classCounts.set(key, (classCounts.get(key) || 0) + 1);
            });
            let bestClass = null;
            let maxC = 0;
            for (let [k, v] of classCounts.entries()) {
                if (v > maxC && v >= 2) { maxC = v; bestClass = k; }
            }
            if (bestClass && maxC >= 2) {
                cards = Array.from(new Set(candidates.filter(c => (c.tagName + "|" + c.className) === bestClass)));
                data.product_detection_method = "Structural repeated containers";
            }
        }
    }

    // Deduplicate cards
    let seenCards = new Set();
    cards.forEach(c => {
        const linkList = Array.from(c.querySelectorAll('a[href]'));
        let link = linkList.find(l => l.innerText.trim().length > 5 && !l.href.includes('javascript:'));
        if (!link) link = linkList[0];
        
        const title = link && link.innerText.trim() ? link.innerText.trim().replace(/\n/g, ' ') : c.innerText.trim().split('\n').find(l => l.trim().length > 0) || "";
        const url = (link && link.href) ? link.href : title;
        const key = url || title;
        if (key && !seenCards.has(key)) {
            seenCards.add(key);
            data.rendered_product_cards++;
            if (data.sample_products.length < 3) {
                data.sample_products.push(String(title).substring(0, 50));
            }
        }
    });
    
    let productBottom = 0;
    if (productRoot && productRoot !== document.body && isVisible(productRoot)) {
        productBottom = productRoot.getBoundingClientRect().bottom + window.scrollY;
    } else if (cards.length > 0) {
        const lastCard = cards[cards.length - 1];
        productBottom = lastCard.getBoundingClientRect().bottom + window.scrollY;
    }

    // Explicit zero products
    const zeroRegex = /(申し訳ございません。ただいま該当する製品がみつかりません|大変申し訳ございません。ただいまご購入いただけるモデルがございません|no products found|0 results)/i;
    if (document.body.innerText.match(zeroRegex)) {
        data.audit_notes.push("Explicit zero-product state detected.");
        data.reported_product_count = 0;
        data.rendered_product_cards = 0;
    }

    // 3. Q&A
    const faqs = mainRegion.querySelectorAll('.faqs .sectionLi');
    if (faqs.length > 0 && Array.from(faqs).some(isVisible)) {
        data.qa_available = true;
        data.qa_count = Array.from(faqs).filter(isVisible).length;
        data.qa_detection_method = "Site-specific FAQ selector";
        for (let i = 0; i < Math.min(3, faqs.length); i++) {
            const titleEl = faqs[i].querySelector('h3, .se-sl-ti');
            if (titleEl) data.sample_qa.push(titleEl.innerText.trim());
        }
    } else {
        // Generic fallback
        const accordions = Array.from(mainRegion.querySelectorAll('details, [aria-expanded], [data-accordion]'))
            .filter(isVisible)
            .filter(el => {
                const top = el.getBoundingClientRect().top + window.scrollY;
                if (top > footerTop || el.closest('nav, aside, .filter, header')) return false;
                const text = el.innerText.trim();
                return text.includes('?') || el.closest('.faq, .q-a, .questions, [class*="faq" i], [class*="question" i]');
            });
        
        let parents = new Map();
        accordions.forEach(el => {
            const p = el.parentElement;
            parents.set(p, (parents.get(p) || 0) + 1);
        });
        
        for (let [p, count] of parents.entries()) {
            if (count >= 2) {
                data.qa_available = true;
                data.qa_count = count;
                data.qa_detection_method = "Generic accordion heuristic";
                const items = Array.from(p.querySelectorAll('details summary, [aria-expanded]'));
                for (let i = 0; i < Math.min(3, items.length); i++) {
                    data.sample_qa.push(items[i].innerText.trim());
                }
                break;
            }
        }
    }

    // Component Sequence & BOPC
    const topLevelElements = Array.from(mainRegion.children).flatMap(c => {
        if (c.tagName === 'DIV' && c.children.length === 1) return Array.from(c.children); // unwrap trivial div
        return [c];
    }).filter(isVisible);

    let bopcFound = false;

    topLevelElements.forEach((c, index) => {
        const top = c.getBoundingClientRect().top + window.scrollY;
        const bottom = c.getBoundingClientRect().bottom + window.scrollY;
        
        const compName = c.getAttribute('componentname') || c.getAttribute('componentName') || c.id || c.className || "Block";
        const isFaq = c.querySelector('.faqs') || c.classList.contains('faqs') || (data.qa_available && data.qa_detection_method.includes('generic') && c.querySelector('details, [aria-expanded]'));
        const isProduct = c === productRoot || c.contains(productRoot);
        const text = c.innerText.trim();
        const charCount = text.length;
        
        data.component_sequence.push(`${c.tagName.toLowerCase()}(${String(compName).substring(0,25)})[${charCount} chars]`);

        // BOPC Candidates
        if (bottom <= footerTop + 100 && !isFaq) {
            if (c.tagName === 'NAV' || c.tagName === 'HEADER' || c.tagName === 'FOOTER') return;

            const headings = Array.from(c.querySelectorAll('h1, h2, h3, h4, h5')).filter(isVisible).map(h => h.innerText.trim()).filter(h => h.length > 0);
            const paragraphs = Array.from(c.querySelectorAll('p, .text')).filter(isVisible).filter(p => p.innerText.trim().length > 40);
            
            const debugObj = {
                "DOM Order": index,
                "Component Name": compName,
                "Tag": c.tagName,
                "Text Characters": charCount,
                "Heading Count": headings.length,
                "Paragraph Count": paragraphs.length,
                "Contains FAQ": !!isFaq,
                "Contains Products": !!isProduct,
                "Vertical Pos": `${Math.round(top)} - ${Math.round(bottom)}`,
                "BOPC Yes/No": "No"
            };

            if (!bopcFound && charCount > 300 && headings.length >= 1 && paragraphs.length >= 1) {
                if (!isProduct || (charCount > 1500 && paragraphs.length >= 3)) {
                    bopcFound = true;
                    debugObj["BOPC Yes/No"] = "Yes";
                    data.bopc_available = true;
                    data.bopc_char_count = charCount;
                    data.bopc_heading_count = headings.length;
                    data.bopc_headings = headings.slice(0, 3);
                    
                    const isLenovoBOPC = (typeof compName === 'string' && (compName.includes('ofp-Bottom-Content') || compName.includes('htmlUpload')));
                    data.bopc_detection_method = isLenovoBOPC ? "Site-specific long-form block" : "Generic structural long-form block";
                    data.component_sequence[data.component_sequence.length - 1] += " -> BOPC!";
                } else {
                    debugObj["Excluded Reason"] = "Container is primarily product grid";
                }
            } else if (!bopcFound) {
                debugObj["Excluded Reason"] = "Thresholds not met (chars>300, headings>=1, paragraphs>=1)";
            }
            data.debug_candidates.push(debugObj);
        }
    });

    return data;
}
"""

def assign_page_status(data):
    if data.get("http_status", "") == "Error":
        return "Manual Review"
    
    notes = " ".join(data["audit_notes"])
    
    if "Explicit zero-product" in notes or (data["reported_product_count"] == 0 and data["rendered_product_cards"] == 0 and data["product_region_found"]):
        return "Missing Products"
        
    if data["product_region_found"] and data["rendered_product_cards"] == 0:
        return "Products Not Loading"
        
    if not data["product_region_found"] and not data["bopc_available"] and not data["qa_available"]:
        return "Missing Content"
        
    if data["rendered_product_cards"] > 0:
        return "OK"
        
    return "Manual Review"

def ensure_playwright_installed():
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            browser.close()
    except PlaywrightError as e:
        if "Executable doesn't exist" in str(e) or "playwright install" in str(e):
            import subprocess
            subprocess.run(["playwright", "install", "chromium"], check=True)
        else:
            raise e

def process_single_url(url: str) -> dict:
    result = {
        "Address": url,
        "Input Status Code": 200,
        "HTTP Status": "Error",
        "Final URL": url,
        "Page Heading": "",
        "Product Region Found": False,
        "Reported Product Count": 0,
        "Rendered Product Cards": 0,
        "Sample Product Names": "",
        "Product Loading/Skeleton Remaining": False,
        "Product Detection Method": "",
        "Is BOPC Available?": "No",
        "BOPC Character Count": 0,
        "BOPC Heading Count": 0,
        "BOPC Headings": "",
        "BOPC Detection Method": "",
        "Is Q&A Available?": "No",
        "Q&A Count": 0,
        "Sample Q&A Questions": "",
        "Q&A Detection Method": "",
        "Page Status": "Manual Review",
        "Audit Evidence": "",
        "Component Sequence": "",
        "Checked At": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context()
        page = context.new_page()
        try:
            response = page.goto(url, wait_until="domcontentloaded", timeout=45000)
            if response:
                result["HTTP Status"] = f"{response.status} {response.status_text}"
            result["Final URL"] = page.url
            
            # Fast scroll till bottom or stable
            page.evaluate('''
                () => new Promise(resolve => {
                    let stableChecks = 0;
                    let prevHeight = 0;
                    let attempts = 0;
                    const interval = setInterval(() => {
                        window.scrollBy(0, 800);
                        const h = document.body.scrollHeight;
                        const skeletons = Array.from(document.querySelectorAll('.skeleton, .skeleton_product_card, .pc_dlp_skeleton_box'));
                        const skeletonVisible = skeletons.some(el => {
                            const style = window.getComputedStyle(el);
                            return style.display !== 'none' && style.visibility !== 'hidden' && style.opacity !== '0';
                        });
                        if (h === prevHeight && !skeletonVisible) {
                            stableChecks++;
                            if (stableChecks >= 3) {
                                clearInterval(interval);
                                resolve();
                            }
                        } else {
                            stableChecks = 0;
                        }
                        prevHeight = h;
                        attempts++;
                        if (attempts >= 40) { // max ~10s
                            clearInterval(interval);
                            resolve();
                        }
                    }, 250);
                })
            ''')
            
            data = page.evaluate(JS_EXTRACT_SCRIPT)
            
            # Retry logic: DO NOT erase first result if reload fails
            if data["product_region_found"] and data["rendered_product_cards"] == 0 and data["reported_product_count"] > 0:
                page.wait_for_timeout(3000)
                data2 = page.evaluate(JS_EXTRACT_SCRIPT)
                if data2["rendered_product_cards"] == 0:
                    try:
                        page.reload(wait_until="domcontentloaded", timeout=20000)
                        page.evaluate('''
                            () => new Promise(resolve => {
                                let attempts = 0;
                                const interval = setInterval(() => {
                                    window.scrollBy(0, 800);
                                    if (++attempts >= 30) {
                                        clearInterval(interval);
                                        resolve();
                                    }
                                }, 300);
                            })
                        ''')
                        data3 = page.evaluate(JS_EXTRACT_SCRIPT)
                        if data3["rendered_product_cards"] > 0 or not data3["product_loading_skeleton"]:
                            data = data3
                        else:
                            data["audit_notes"].append("Reload retry completed but still suspicious.")
                    except Exception as e:
                        data["audit_notes"].append("Reload retry timed out; retained first rendered-DOM result.")
                else:
                    data = data2
            
            data["http_status"] = result["HTTP Status"]
            status = assign_page_status(data)
            
            # Map back
            result["Page Heading"] = data["page_heading"]
            result["Product Region Found"] = data["product_region_found"]
            result["Reported Product Count"] = data["reported_product_count"]
            result["Rendered Product Cards"] = data["rendered_product_cards"]
            result["Sample Product Names"] = " | ".join(data["sample_products"])
            result["Product Loading/Skeleton Remaining"] = data["product_loading_skeleton"]
            result["Product Detection Method"] = data["product_detection_method"]
            result["Is BOPC Available?"] = "Yes" if data["bopc_available"] else "No"
            result["BOPC Character Count"] = data["bopc_char_count"]
            result["BOPC Heading Count"] = data["bopc_heading_count"]
            result["BOPC Headings"] = " | ".join(data["bopc_headings"])
            result["BOPC Detection Method"] = data["bopc_detection_method"]
            result["Is Q&A Available?"] = "Yes" if data["qa_available"] else "No"
            result["Q&A Count"] = data["qa_count"]
            result["Sample Q&A Questions"] = " | ".join(data["sample_qa"])
            result["Q&A Detection Method"] = data["qa_detection_method"]
            result["Page Status"] = status
            
            # Combine audit notes and debug strings
            audit_str = " ".join(data["audit_notes"])
            if data.get("debug_candidates"):
                audit_str += " | BOPC Candidates: " + " ; ".join([f"[{c['Component Name']}] chars={c['Text Characters']} headings={c['Heading Count']} -> {c['Excluded Reason'] if 'Excluded Reason' in c else 'BOPC!'}" for c in data["debug_candidates"]])
            
            result["Audit Evidence"] = audit_str
            result["Component Sequence"] = " > ".join(data["component_sequence"])
            
        except Exception as e:
            result["Audit Evidence"] = f"Browser Error: {str(e).splitlines()[0]}"
            result["Page Status"] = "Manual Review"
        finally:
            context.close()
            browser.close()
            
    return result

def generate_markdown(results):
    lines = []
    for r in results:
        lines.append(f"## {r['Address']}")
        lines.append("\n### Page")
        lines.append(f"- HTTP Status: {r['HTTP Status']}")
        lines.append(f"- Final URL: {r['Final URL']}")
        lines.append(f"- Page Heading: {r['Page Heading']}")
        lines.append(f"- Page Status: {r['Page Status']}")
        
        lines.append("\n### Products")
        lines.append(f"- Product Region Found: {r['Product Region Found']}")
        lines.append(f"- Reported Product Count: {r['Reported Product Count']}")
        lines.append(f"- Rendered Product Cards: {r['Rendered Product Cards']}")
        lines.append(f"- Product Detection Method: {r['Product Detection Method']}")
        lines.append(f"- Loading/Skeleton Remaining: {r['Product Loading/Skeleton Remaining']}")
        lines.append(f"- Sample Products: {r['Sample Product Names']}")
        
        lines.append("\n### Q&A")
        lines.append(f"- Available: {r['Is Q&A Available?']}")
        lines.append(f"- Detection Method: {r['Q&A Detection Method']}")
        lines.append(f"- Q&A Count: {r['Q&A Count']}")
        lines.append(f"- Sample Questions: {r['Sample Q&A Questions']}")
        
        lines.append("\n### BOPC")
        lines.append(f"- Available: {r['Is BOPC Available?']}")
        lines.append(f"- Detection Method: {r['BOPC Detection Method']}")
        lines.append(f"- Character Count: {r['BOPC Character Count']}")
        lines.append(f"- Heading Count: {r['BOPC Heading Count']}")
        lines.append(f"- Headings: {r['BOPC Headings']}")
        
        lines.append("\n### Structural Evidence")
        lines.append(f"- Component Sequence: {r['Component Sequence']}")
        lines.append(f"- Audit Notes: {r['Audit Evidence']}")
        lines.append("\n---\n")
    return "\n".join(lines)

def generate_excel(results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DLP Content Audit"
    
    if not results:
        return b""
        
    headers = list(results[0].keys())
    ws.append(headers)
    
    # Fill colors
    green_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    red_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    orange_fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
    gray_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")
    
    for row_idx, r in enumerate(results, start=2):
        for col_idx, key in enumerate(headers, start=1):
            val = r[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            
            if key in ["Is BOPC Available?", "Is Q&A Available?"]:
                if val == "Yes": cell.fill = green_fill
                else: cell.fill = gray_fill
                
            if key == "Page Status":
                if val == "OK": cell.fill = green_fill
                elif val in ["Missing Products", "Products Not Loading", "Missing Content"]: cell.fill = red_fill
                elif val == "Manual Review": cell.fill = orange_fill
                
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    
    # Summary sheet
    ws2 = wb.create_sheet("Audit Summary")
    summary_data = {
        "Total URLs Processed": len(results),
        "Successful Browser Loads": sum(1 for r in results if "Error" not in r["HTTP Status"]),
        "Errors": sum(1 for r in results if "Error" in r["HTTP Status"]),
        "BOPC Yes": sum(1 for r in results if r["Is BOPC Available?"] == "Yes"),
        "BOPC No": sum(1 for r in results if r["Is BOPC Available?"] == "No"),
        "Q&A Yes": sum(1 for r in results if r["Is Q&A Available?"] == "Yes"),
        "Q&A No": sum(1 for r in results if r["Is Q&A Available?"] == "No"),
        "OK": sum(1 for r in results if r["Page Status"] == "OK"),
        "Missing Content": sum(1 for r in results if r["Page Status"] == "Missing Content"),
        "Missing Products": sum(1 for r in results if r["Page Status"] == "Missing Products"),
        "Products Not Loading": sum(1 for r in results if r["Page Status"] == "Products Not Loading"),
        "Manual Review": sum(1 for r in results if r["Page Status"] == "Manual Review")
    }
    
    for k, v in summary_data.items():
        ws2.append([k, v])
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def dlp_content_audit_tool():
    st.title("Category/DLP Content & Product Audit")
    st.caption("Audit rendered DLP pages for products, Q&A, bottom-of-page content (BOPC), and page-loading issues.")
    
    mode = st.radio("Input Mode", ["Single URL Test", "Paste URL List", "Bulk Excel Upload"], horizontal=True)
    
    if mode == "Single URL Test":
        url = st.text_input("Enter URL to audit")
        if st.button("Run DLP Test", type="primary"):
            ensure_playwright_installed()
            with st.spinner("Loading and auditing page..."):
                res = process_single_url(url)
                
                c1, c2, c3 = st.columns(3)
                c1.metric("Page Status", res["Page Status"])
                c2.metric("HTTP Status", res["HTTP Status"])
                c3.metric("Final URL", res["Final URL"])
                
                st.write(f"**Page Heading:** {res['Page Heading']}")
                
                c1, c2 = st.columns(2)
                c1.metric("Reported Product Count", res["Reported Product Count"])
                c2.metric("Rendered Product Cards", res["Rendered Product Cards"])
                
                c1, c2 = st.columns(2)
                c1.metric("Is Q&A Available?", res["Is Q&A Available?"])
                c2.metric("Q&A Count", res["Q&A Count"])
                
                c1, c2 = st.columns(2)
                c1.metric("Is BOPC Available?", res["Is BOPC Available?"])
                c2.metric("BOPC Character Count", res["BOPC Character Count"])
                
                with st.expander("Product evidence"):
                    st.write(f"Region Found: {res['Product Region Found']}")
                    st.write(f"Detection Method: {res['Product Detection Method']}")
                    st.write(f"Skeleton Remaining: {res['Product Loading/Skeleton Remaining']}")
                    st.write(f"Sample Products: {res['Sample Product Names']}")
                    
                with st.expander("Q&A questions"):
                    st.write(f"Detection Method: {res['Q&A Detection Method']}")
                    st.write(f"Sample Q&A: {res['Sample Q&A Questions']}")
                    
                with st.expander("BOPC headings"):
                    st.write(f"Heading Count: {res['BOPC Heading Count']}")
                    st.write(f"Detection Method: {res['BOPC Detection Method']}")
                    st.write(f"Headings: {res['BOPC Headings']}")
                    
                with st.expander("Component sequence"):
                    st.write(res["Component Sequence"])
                    
                with st.expander("Audit notes"):
                    st.write(res["Audit Evidence"])
                    
    elif mode == "Paste URL List":
        urls_text = st.text_area("Paste URLs (one per line)")
        workers = st.slider("Concurrency", 1, 3, 2)
        if st.button("Run Bulk Audit", type="primary"):
            urls = [u.strip() for u in urls_text.splitlines() if u.strip()]
            run_bulk_audit(urls, workers)
            
    elif mode == "Bulk Excel Upload":
        uploaded = st.file_uploader("Upload Excel file", type=["xlsx"])
        workers = st.slider("Concurrency", 1, 3, 2)
        if uploaded and st.button("Run Bulk Audit", type="primary"):
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.getvalue()))
            if "Pages" in wb.sheetnames:
                sheet = wb["Pages"]
                urls = []
                for row_idx in range(2, sheet.max_row + 1):
                    url = sheet.cell(row=row_idx, column=1).value
                    status_code = sheet.cell(row=row_idx, column=2).value
                    if url and str(status_code).strip() == "200":
                        urls.append(url.strip())
                run_bulk_audit(urls, workers)
            else:
                st.error("Excel file must contain a 'Pages' sheet.")
                
def run_bulk_audit(urls, workers):
    if not urls:
        st.warning("No valid URLs to process.")
        return
        
    ensure_playwright_installed()
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    results = []
    
    # Process with separate contexts per thread
    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(process_single_url, url): url for url in urls}
        done_count = 0
        for future in concurrent.futures.as_completed(futures):
            res = future.result()
            results.append(res)
            done_count += 1
            progress_bar.progress(done_count / len(urls))
            status_text.text(f"Processed {done_count}/{len(urls)} URLs...")
            
    st.session_state["dlp_audit_results"] = results
    display_bulk_results()
    
def display_bulk_results():
    if "dlp_audit_results" in st.session_state:
        results = st.session_state["dlp_audit_results"]
        st.success("Audit complete!")
        
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("URLs Processed", len(results))
        c2.metric("BOPC Yes", sum(1 for r in results if r["Is BOPC Available?"] == "Yes"))
        c3.metric("Q&A Yes", sum(1 for r in results if r["Is Q&A Available?"] == "Yes"))
        c4.metric("OK", sum(1 for r in results if r["Page Status"] == "OK"))
        c5.metric("Issues/Manual", sum(1 for r in results if r["Page Status"] != "OK"))
        
        excel_data = generate_excel(results)
        md_data = generate_markdown(results)
        
        c1, c2 = st.columns(2)
        c1.download_button("Download Excel Audit (.xlsx)", excel_data, "DLP_Content_Audit.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        c2.download_button("Download Audit Evidence (.md)", md_data, "DLP_Audit_Evidence.md", "text/markdown", use_container_width=True)
