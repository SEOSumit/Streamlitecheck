from __future__ import annotations

import html as html_lib
import re
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urljoin, urlsplit, urlunsplit
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/150.0.0.0 Safari/537.36"
)


@dataclass(frozen=True)
class LinkItem:
    row: int
    page_url: str
    paragraph: str
    anchor_text: str
    suggested_url: str


@dataclass
class PageResult:
    requested_url: str
    final_url: str = ""
    status_code: int | None = None
    body: str = ""
    error: str = ""


@dataclass
class ParsedAnchor:
    href: str
    text_parts: list[str]


@dataclass
class ParsedContainer:
    tag: str
    text_parts: list[str]
    anchors: list[ParsedAnchor]


class ParagraphHTMLParser(HTMLParser):
    """Collect text and links from paragraph-like HTML containers."""

    CONTAINER_TAGS = {"p", "li", "dd", "td", "div"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.containers: list[ParsedContainer] = []
        self._open_containers: list[ParsedContainer] = []
        self._open_anchors: list[ParsedAnchor] = []
        self._ignored_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.lower()
        if tag in {"script", "style", "noscript"}:
            self._ignored_depth += 1
            return
        if self._ignored_depth:
            return
        if tag in self.CONTAINER_TAGS:
            container = ParsedContainer(tag, [], [])
            self.containers.append(container)
            self._open_containers.append(container)
        if tag == "a":
            href = dict(attrs).get("href", "") or ""
            anchor = ParsedAnchor(str(href), [])
            self._open_anchors.append(anchor)
            for container in self._open_containers:
                container.anchors.append(anchor)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"script", "style", "noscript"}:
            self._ignored_depth = max(0, self._ignored_depth - 1)
            return
        if self._ignored_depth:
            return
        if tag == "a" and self._open_anchors:
            self._open_anchors.pop()
        if tag in self.CONTAINER_TAGS:
            for index in range(len(self._open_containers) - 1, -1, -1):
                if self._open_containers[index].tag == tag:
                    del self._open_containers[index]
                    break

    def handle_data(self, data: str) -> None:
        if self._ignored_depth:
            return
        for container in self._open_containers:
            container.text_parts.append(data)
        for anchor in self._open_anchors:
            anchor.text_parts.append(data)


def normalize_text(value: object) -> str:
    text = html_lib.unescape(str(value or ""))
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\u200b", "").replace("\ufeff", "")
    return re.sub(r"\s+", " ", text).strip().casefold()


def normalize_url(value: str, base_url: str = "") -> str:
    absolute = urljoin(base_url, str(value or "").strip())
    parts = urlsplit(absolute)
    scheme = parts.scheme.lower()
    hostname = (parts.hostname or "").lower()
    port = parts.port
    if port and not ((scheme == "http" and port == 80) or (scheme == "https" and port == 443)):
        hostname = f"{hostname}:{port}"
    path = re.sub(r"/{2,}", "/", parts.path or "/")
    if path != "/":
        path = path.rstrip("/")
    return urlunsplit((scheme, hostname, path, parts.query, ""))


def extract_items(ws) -> list[LinkItem]:
    items: list[LinkItem] = []
    current_page = ""

    # Row 1 contains the workbook headers in the supplied template.
    for row in range(2, ws.max_row + 1):
        paragraph = ws.cell(row, 3).value
        anchor = ws.cell(row, 4).value
        suggested = ws.cell(row, 5).value

        if isinstance(paragraph, str) and paragraph.strip().lower().startswith(("http://", "https://")):
            if not anchor and not suggested:
                current_page = paragraph.strip()
                continue

        if paragraph and anchor and suggested:
            if not current_page:
                raise ValueError(
                    f"Row {row}: source page URL is missing above this suggestion. "
                    "Expected the page URL in column C."
                )
            items.append(
                LinkItem(
                    row=row,
                    page_url=current_page,
                    paragraph=str(paragraph).strip(),
                    anchor_text=str(anchor).strip(),
                    suggested_url=str(suggested).strip(),
                )
            )

    if not items:
        raise ValueError(
            "No checkable rows found. Expected page URLs and linking paragraphs in column C, "
            "anchor text in D, and suggested hyperlinks in E."
        )
    return items


def fetch_page(url: str, timeout: int = 30) -> PageResult:
    try:
        request = Request(
            url,
            headers={
                "User-Agent": USER_AGENT,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
                "Cache-Control": "no-cache",
            },
        )
        with urlopen(request, timeout=timeout) as response:
            status = response.status
            charset = response.headers.get_content_charset() or "utf-8"
            body = response.read(25_000_000).decode(charset, errors="replace")
            result = PageResult(
                requested_url=url,
                final_url=response.geturl(),
                status_code=status,
                body=body,
            )
            if not (200 <= status < 300):
                result.error = f"Page returned HTTP {status}"
            return result
    except HTTPError as exc:
        return PageResult(
            requested_url=url,
            final_url=exc.geturl(),
            status_code=exc.code,
            error=f"Page returned HTTP {exc.code}",
        )
    except (URLError, TimeoutError, OSError) as exc:
        return PageResult(requested_url=url, error=f"Page request failed: {exc}")


def fetch_pages(
    urls: Iterable[str],
    max_workers: int = 5,
    progress: Callable[[int, int], None] | None = None,
) -> dict[str, PageResult]:
    unique_urls = list(dict.fromkeys(urls))
    completed = 0
    results: dict[str, PageResult] = {}
    with ThreadPoolExecutor(max_workers=max(1, min(max_workers, 10))) as pool:
        futures = {pool.submit(fetch_page, url): url for url in unique_urls}
        for future in as_completed(futures):
            url = futures[future]
            try:
                results[url] = future.result()
            except Exception as exc:  # defensive: one page must not stop the workbook
                results[url] = PageResult(url, error=f"Unexpected page error: {exc}")
            completed += 1
            if progress:
                progress(completed, len(unique_urls))
    return results


def _matching_containers(parsed: ParagraphHTMLParser, paragraph: str):
    expected = normalize_text(paragraph)
    matches = []
    for container in parsed.containers:
        container_text = normalize_text("".join(container.text_parts))
        if expected and expected in container_text:
            matches.append((len(container_text), container))
    matches.sort(key=lambda pair: pair[0])
    return [tag for _, tag in matches]


def check_item(item: LinkItem, page: PageResult) -> dict[str, object]:
    base = {
        "Excel Row": item.row,
        "Source Page URL": item.page_url,
        "Linking Paragraph": item.paragraph,
        "Suggested Anchor Text": item.anchor_text,
        "Suggested Hyperlink": item.suggested_url,
        "Implementation Status": "Not Implemented",
        "Actual Hyperlink Found": "",
        "Check Details": "",
        "Page HTTP Status": page.status_code or "",
        "Final Page URL": page.final_url,
    }

    if page.error:
        base["Check Details"] = page.error
        return base

    parsed = ParagraphHTMLParser()
    parsed.feed(page.body)
    containers = _matching_containers(parsed, item.paragraph)
    if not containers:
        base["Check Details"] = "Specified paragraph was not found on the page"
        return base

    expected_anchor = normalize_text(item.anchor_text)
    expected_url = normalize_url(item.suggested_url, item.page_url)
    actual_urls: list[str] = []

    for container in containers:
        for link in container.anchors:
            if normalize_text("".join(link.text_parts)) != expected_anchor:
                continue
            href = link.href.strip()
            if href:
                resolved = urljoin(page.final_url or item.page_url, href)
                if resolved not in actual_urls:
                    actual_urls.append(resolved)
                if normalize_url(resolved) == expected_url:
                    base["Implementation Status"] = "Implemented"
                    base["Actual Hyperlink Found"] = resolved
                    base["Check Details"] = "Anchor and suggested hyperlink match within the specified paragraph"
                    return base

    if actual_urls:
        base["Actual Hyperlink Found"] = " | ".join(actual_urls)
        base["Check Details"] = "Anchor is linked within the paragraph, but the destination URL is different"
    else:
        base["Check Details"] = "Paragraph was found, but the specified anchor text is not linked within it"
    return base


def _result_columns(ws) -> dict[str, int]:
    desired = [
        "Source Page URL",
        "Implementation Status",
        "Actual Hyperlink Found",
        "Check Details",
        "Page HTTP Status",
        "Final Page URL",
        "Checked At",
    ]
    existing = {
        str(ws.cell(1, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value
    }
    next_col = ws.max_column + 1
    mapping: dict[str, int] = {}
    for header in desired:
        if header in existing:
            mapping[header] = existing[header]
        else:
            mapping[header] = next_col
            next_col += 1
    return mapping


def write_results(ws, results: list[dict[str, object]], checked_at: str) -> None:
    columns = _result_columns(ws)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    green_font = Font(name="Calibri", size=10, bold=True, color="006100")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    red_font = Font(name="Calibri", size=10, bold=True, color="9C0006")

    for header, col in columns.items():
        cell = ws.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "Source Page URL": 55,
        "Implementation Status": 20,
        "Actual Hyperlink Found": 55,
        "Check Details": 62,
        "Page HTTP Status": 18,
        "Final Page URL": 55,
        "Checked At": 23,
    }
    for header, width in widths.items():
        ws.column_dimensions[get_column_letter(columns[header])].width = width

    for result in results:
        row = int(result["Excel Row"])
        for header in columns:
            value = checked_at if header == "Checked At" else result.get(header, "")
            cell = ws.cell(row, columns[header], value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        status_cell = ws.cell(row, columns["Implementation Status"])
        if result["Implementation Status"] == "Implemented":
            status_cell.fill = green_fill
            status_cell.font = green_font
        else:
            status_cell.fill = red_fill
            status_cell.font = red_font

    ws.auto_filter.ref = f"A1:{get_column_letter(max(columns.values()))}{ws.max_row}"
    if not ws.freeze_panes:
        ws.freeze_panes = "A2"


def add_summary_sheet(wb, results: list[dict[str, object]], checked_at: str) -> None:
    title = "Link Check Summary"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title, 0)

    implemented = sum(r["Implementation Status"] == "Implemented" for r in results)
    not_implemented = len(results) - implemented
    unique_pages = len({r["Source Page URL"] for r in results})
    fetch_errors = sum(
        str(r["Check Details"]).startswith(("Page request failed", "Page returned HTTP", "Unexpected page error"))
        for r in results
    )

    rows = [
        ("Internal Link Implementation Check", ""),
        ("Checked At", checked_at),
        ("Source Pages", unique_pages),
        ("Suggestions Checked", len(results)),
        ("Implemented", implemented),
        ("Not Implemented", not_implemented),
        ("Rows with Page Fetch Errors", fetch_errors),
        ("Rule", "Implemented only when the specified paragraph contains the exact anchor text linked to the suggested URL."),
    ]
    for row in rows:
        ws.append(row)

    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A1:B1")
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 1).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor="D9EAF7")
        ws.cell(row, 1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, 2).alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 31
    ws.column_dimensions["B"].width = 92
    ws.row_dimensions[1].height = 25


def process_workbook(
    workbook_bytes: bytes,
    sheet_name: str,
    max_workers: int = 5,
    progress: Callable[[int, int], None] | None = None,
) -> tuple[bytes, list[dict[str, object]]]:
    wb = load_workbook(BytesIO(workbook_bytes))
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    items = extract_items(ws)
    pages = fetch_pages((item.page_url for item in items), max_workers=max_workers, progress=progress)
    results = [check_item(item, pages[item.page_url]) for item in items]
    checked_at = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y-%m-%d %H:%M:%S IST")
    write_results(ws, results, checked_at)
    add_summary_sheet(wb, results, checked_at)
    output = BytesIO()
    wb.save(output)
    return output.getvalue(), results


def workbook_sheet_names(workbook_bytes: bytes) -> list[str]:
    wb = load_workbook(BytesIO(workbook_bytes), read_only=True)
    return wb.sheetnames


def output_filename(original_name: str) -> str:
    stem = Path(original_name).stem
    return f"{stem}-link-check-results.xlsx"
