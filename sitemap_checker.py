from __future__ import annotations

import csv
import html as html_lib
import os
import re
import time
import unicodedata
import zipfile
from copy import copy
from dataclasses import dataclass
from html.parser import HTMLParser
from io import BytesIO, StringIO
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import unquote, urljoin, urlsplit, urlunsplit
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/150.0.0.0 Safari/537.36"
)

GEMINI_MODEL_NAME = "gemini-1.5-flash"


@dataclass
class DataTable:
    name: str
    rows: list[list[object]]


class SitemapAnchorParser(HTMLParser):
    def __init__(self, base_url: str) -> None:
        super().__init__(convert_charrefs=True)
        self.base_url = base_url
        self.anchors: list[dict[str, str]] = []
        self._open_anchor: dict[str, object] | None = None
        self._ignored_depth = 0

    def handle_starttag(self, tag: str, attrs) -> None:
        tag = tag.casefold()
        attrs_dict = dict(attrs)
        if tag in {"script", "style", "noscript"}:
            self._ignored_depth += 1
            return
        if self._ignored_depth:
            return
        if tag == "base" and attrs_dict.get("href"):
            self.base_url = urljoin(self.base_url, str(attrs_dict["href"]))
        elif tag == "a":
            self._open_anchor = {
                "href": str(attrs_dict.get("href") or "").strip(),
                "text": [],
            }
        elif tag == "img" and self._open_anchor is not None and attrs_dict.get("alt"):
            self._open_anchor["text"].append(str(attrs_dict["alt"]))

    def handle_endtag(self, tag: str) -> None:
        tag = tag.casefold()
        if tag in {"script", "style", "noscript"}:
            self._ignored_depth = max(0, self._ignored_depth - 1)
            return
        if self._ignored_depth:
            return
        if tag == "a" and self._open_anchor is not None:
            href = str(self._open_anchor["href"])
            if href and not href.casefold().startswith(("mailto:", "tel:", "javascript:")):
                absolute = urljoin(self.base_url, href)
                if absolute.startswith(("http://", "https://")):
                    parts = urlsplit(absolute)
                    absolute = urlunsplit((parts.scheme, parts.netloc, parts.path, parts.query, ""))
                    self.anchors.append(
                        {
                            "url": absolute,
                            "anchor": normalize_text("".join(self._open_anchor["text"])),
                        }
                    )
            self._open_anchor = None

    def handle_data(self, data: str) -> None:
        if not self._ignored_depth and self._open_anchor is not None:
            self._open_anchor["text"].append(data)


def normalize_text(value: object) -> str:
    text = html_lib.unescape(str(value or ""))
    text = unicodedata.normalize("NFKC", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_compare_url(value: str) -> str:
    parts = urlsplit(str(value or "").strip())
    scheme = parts.scheme.casefold()
    host = (parts.hostname or "").casefold()
    port = parts.port
    if port and not ((scheme == "http" and port == 80) or (scheme == "https" and port == 443)):
        host = f"{host}:{port}"
    path = re.sub(r"/{2,}", "/", unquote(parts.path or "/"))
    if path != "/":
        path = path.rstrip("/")
    return urlunsplit((scheme, host, path, "", ""))


def clean_url_without_parameters(value: str) -> str:
    parts = urlsplit(str(value).strip())
    return urlunsplit((parts.scheme, parts.netloc, parts.path, "", ""))


def _is_url(value: object) -> bool:
    return isinstance(value, str) and value.strip().casefold().startswith(("http://", "https://"))


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _csv_table(data: bytes, name: str) -> DataTable:
    text = _decode_text(data)
    try:
        dialect = csv.Sniffer().sniff(text[:5000], delimiters=",;\t|")
        rows = [list(row) for row in csv.reader(StringIO(text), dialect)]
    except csv.Error:
        rows = [list(row) for row in csv.reader(StringIO(text))]
    return DataTable(name, rows)


def _xlsx_tables(data: bytes, name: str) -> list[DataTable]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    return [
        DataTable(f"{name} — {ws.title}", [list(row) for row in ws.iter_rows(values_only=True)])
        for ws in wb.worksheets
    ]


def load_tables(data: bytes, filename: str) -> list[DataTable]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".xlsx":
        return _xlsx_tables(data, filename)
    if suffix in {".csv", ".tsv", ".txt"}:
        return [_csv_table(data, filename)]
    if suffix == ".zip":
        tables: list[DataTable] = []
        with zipfile.ZipFile(BytesIO(data)) as archive:
            for member in archive.infolist():
                if member.is_dir() or member.file_size > 50_000_000:
                    continue
                member_data = archive.read(member)
                member_suffix = Path(member.filename).suffix.casefold()
                label = f"{filename}/{member.filename}"
                if member_suffix == ".xlsx":
                    tables.extend(_xlsx_tables(member_data, label))
                elif member_suffix in {".csv", ".tsv", ".txt"}:
                    tables.append(_csv_table(member_data, label))
        if tables:
            return tables
    raise ValueError("Upload an XLSX, CSV, or ZIP containing the URL and status-code list.")


def _clean_status(value: object) -> object:
    if value in (None, ""):
        return ""
    text = str(value).strip()
    match = re.search(r"\b([1-5]\d{2})\b", text)
    return int(match.group(1)) if match else text


def _records_from_table(table: DataTable) -> list[dict[str, object]]:
    records = []
    for row in table.rows[1:]:
        if not row or not _is_url(row[0]):
            continue
        url = str(row[0]).strip()
        records.append(
            {
                "url": url,
                "normalized": normalize_compare_url(url),
                "status": _clean_status(row[1] if len(row) > 1 else ""),
                "final_redirect_url": str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "") else "",
                "has_parameters": bool(urlsplit(url).query),
            }
        )
    return records


def load_url_status_records(data: bytes, filename: str) -> list[dict[str, object]]:
    tables = load_tables(data, filename)
    candidates = []
    
    # Try to find a sheet named 'Pages' specifically
    pages_candidates = [_records_from_table(table) for table in tables if "pages" in table.name.casefold()]
    pages_candidates = [records for records in pages_candidates if records]
    
    if pages_candidates:
        source_records = max(pages_candidates, key=len)
    else:
        candidates = [_records_from_table(table) for table in tables]
        candidates = [records for records in candidates if records]
        if not candidates:
            raise ValueError("No URLs found from A2 onward. Column A must contain URLs and column B status codes.")
        source_records = max(candidates, key=len)

    # Treat clean and parameterized versions as the same URL. Prefer a clean URL when both exist.
    chosen: dict[str, dict[str, object]] = {}
    order: list[str] = []
    for record in source_records:
        key = str(record["normalized"])
        if key not in chosen:
            chosen[key] = record
            order.append(key)
        elif chosen[key]["has_parameters"] and not record["has_parameters"]:
            previous = chosen[key]
            replacement = dict(record)
            if replacement["status"] in (None, ""):
                replacement["status"] = previous["status"]
            if not replacement.get("final_redirect_url"):
                replacement["final_redirect_url"] = previous.get("final_redirect_url", "")
            chosen[key] = replacement
        else:
            if chosen[key]["status"] in (None, "") and record["status"] not in (None, ""):
                chosen[key]["status"] = record["status"]
            if not chosen[key].get("final_redirect_url") and record.get("final_redirect_url"):
                chosen[key]["final_redirect_url"] = record["final_redirect_url"]
    return [chosen[key] for key in order]


def load_queries_from_workbook(data: bytes, filename: str) -> list[str]:
    """Read the optional 'Query' sheet (A1 heading, queries from A2 down) from the same upload."""
    if Path(filename).suffix.casefold() != ".xlsx":
        return []
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return []
    target = None
    for ws in wb.worksheets:
        if ws.title.strip().casefold() in ("query", "queries"):
            target = ws
            break
    if target is None:
        return []
    queries: list[str] = []
    seen: set[str] = set()
    for row in target.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        text = normalize_text(row[0])
        if text and text.casefold() not in seen:
            seen.add(text.casefold())
            queries.append(text)
    return queries


def fetch_sitemap_html(sitemap_url: str, timeout: int = 30) -> tuple[str, str, int]:
    if not sitemap_url.casefold().startswith(("http://", "https://")):
        raise ValueError("Enter a complete HTML sitemap URL beginning with http:// or https://.")
    request = Request(
        sitemap_url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Cache-Control": "no-cache",
        },
    )
    try:
        with urlopen(request, timeout=timeout) as response:
            charset = response.headers.get_content_charset() or "utf-8"
            source = response.read(25_000_000).decode(charset, errors="replace")
            return source, response.geturl(), int(response.status)
    except HTTPError as exc:
        raise ValueError(
            f"The sitemap page returned HTTP {exc.code}. Use the HTML upload or paste fallback."
        ) from exc
    except (URLError, TimeoutError, OSError) as exc:
        raise ValueError(f"The sitemap page could not be fetched: {exc}") from exc


def parse_sitemap_anchors(source: str, sitemap_url: str) -> list[dict[str, str]]:
    parser = SitemapAnchorParser(sitemap_url)
    parser.feed(source)
    sitemap_host = (urlsplit(sitemap_url).hostname or "").casefold()
    anchors = []
    for item in parser.anchors:
        if (urlsplit(item["url"]).hostname or "").casefold() != sitemap_host:
            continue
        item["normalized"] = normalize_compare_url(item["url"])
        anchors.append(item)
    if not anchors:
        raise ValueError("No same-domain HTML links were found in the sitemap source.")
    return anchors


def parse_rendered_links_text(text: str, sitemap_url: str) -> list[dict[str, str]]:
    """Parse newline URLs or tab-separated URL + anchor text copied from Chrome."""
    sitemap_host = (urlsplit(sitemap_url).hostname or "").casefold()
    links: list[dict[str, str]] = []
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "\t" in line:
            url_value, anchor = line.split("\t", 1)
        else:
            url_value, anchor = line, ""
        absolute = urljoin(sitemap_url, url_value.strip())
        if not absolute.casefold().startswith(("http://", "https://")):
            continue
        if (urlsplit(absolute).hostname or "").casefold() != sitemap_host:
            continue
        parts = urlsplit(absolute)
        url_value = urlunsplit((parts.scheme, parts.netloc, parts.path, parts.query, ""))
        links.append(
            {
                "url": url_value,
                "anchor": normalize_text(anchor),
                "normalized": normalize_compare_url(url_value),
            }
        )
    if not links:
        raise ValueError("No same-domain rendered links were found in the pasted data.")
    return links


def collect_rendered_links_browser(sitemap_url: str, timeout: int = 55) -> list[dict[str, str]]:
    """Open the page in Chromium and run the rendered-DOM equivalent of querySelectorAll('a')."""
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.support.ui import WebDriverWait
    except ImportError as exc:
        raise ValueError("Browser automation dependency is not installed. Use the Chrome paste method.") from exc

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(f"--user-agent={USER_AGENT}")
    if os.path.exists("/usr/bin/chromium"):
        options.binary_location = "/usr/bin/chromium"

    driver = None
    try:
        if os.path.exists("/usr/bin/chromedriver"):
            driver = webdriver.Chrome(service=Service("/usr/bin/chromedriver"), options=options)
        else:
            driver = webdriver.Chrome(options=options)
        driver.set_page_load_timeout(timeout)
        driver.get(sitemap_url)
        WebDriverWait(driver, min(timeout, 35)).until(
            lambda current: current.execute_script("return document.readyState") == "complete"
        )
        previous_height = 0
        for _ in range(6):
            height = int(driver.execute_script("return document.body.scrollHeight || 0"))
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight)")
            time.sleep(0.7)
            if height == previous_height:
                break
            previous_height = height
        raw_links = driver.execute_script(
            """
            return [...document.querySelectorAll('a')].map(a => ({
              url: a.href || '',
              anchor: (a.innerText || a.textContent || '').trim().replace(/\\s+/g, ' ')
            }));
            """
        )
    except Exception as exc:
        raise ValueError(
            f"Rendered browser collection failed: {exc}. Use the Chrome paste method instead."
        ) from exc
    finally:
        if driver is not None:
            driver.quit()

    sitemap_host = (urlsplit(sitemap_url).hostname or "").casefold()
    links = []
    for item in raw_links or []:
        url_value = str(item.get("url") or "").strip()
        if not url_value.casefold().startswith(("http://", "https://")):
            continue
        if (urlsplit(url_value).hostname or "").casefold() != sitemap_host:
            continue
        parts = urlsplit(url_value)
        url_value = urlunsplit((parts.scheme, parts.netloc, parts.path, parts.query, ""))
        links.append(
            {
                "url": url_value,
                "anchor": normalize_text(item.get("anchor", "")),
                "normalized": normalize_compare_url(url_value),
            }
        )
    if not links:
        raise ValueError("The rendered page returned no same-domain links. Use the Chrome paste method.")
    return links


def try_fetch_raw_source(sitemap_url: str) -> str | None:
    try:
        source, _, _ = fetch_sitemap_html(sitemap_url)
        return source
    except ValueError:
        return None





def _in_scope(url: str, mode: str, pattern: str) -> bool:
    if mode == "Complete HTML sitemap audit":
        return True
    return pattern.casefold().strip() in clean_url_without_parameters(url).casefold()


ACRONYMS = {
    "ai": "AI", "hpc": "HPC", "gpu": "GPU", "cpu": "CPU", "nas": "NAS",
    "san": "SAN", "sdi": "SDI", "cio": "CIO", "it": "IT", "vmware": "VMware",
    "truscale": "TruScale", "lenovo": "Lenovo", "thinksystem": "ThinkSystem",
    "thinkagile": "ThinkAgile", "neptune": "Neptune", "amd": "AMD",
}


def suggest_anchor_text(url: str) -> str:
    parts = [part for part in unquote(urlsplit(url).path).split("/") if part]
    ignored = {"in", "en", "us", "uk", "br", "pt", "c", "p", "d", "index", "html"}
    slug = next((part for part in reversed(parts) if part.casefold() not in ignored), "")
    words = [word for word in re.split(r"[-_]+", slug) if word]
    return " ".join(ACRONYMS.get(word.casefold(), word.capitalize()) for word in words)


def suggest_missing_anchor_for_future_ai(url: str, api_key: str | None = None) -> str:
    """Extension point for a future AI anchor-text provider; intentionally blank today."""
    return ""


def _status_group(status: object) -> str:
    try:
        code = int(status)
    except (TypeError, ValueError):
        return "other"
    if code == 200:
        return "200"
    if 300 <= code < 400:
        return "3xx"
    return "other"


def _audit_rows(
    uploaded_records: list[dict[str, object]],
    sitemap_anchors: list[dict[str, str]],
    raw_source: str | None,
    sitemap_url: str,
    mode: str,
    pattern: str,
    queries: list[str] | None = None,
    api_key: str | None = None,
    progress=None,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    scoped_uploaded = [r for r in uploaded_records if _in_scope(str(r["url"]), mode, pattern)]
    scoped_anchors = [r for r in sitemap_anchors if _in_scope(r["url"], mode, pattern)]
    if mode != "Complete HTML sitemap audit" and not pattern.strip():
        raise ValueError("Enter a folder or URL pattern, for example /servers-storage/.")

    uploaded_map = {str(record["normalized"]): record for record in scoped_uploaded}
    sitemap_keys = {item["normalized"] for item in scoped_anchors}

    raw_vanilla_keys: set[str] | None = None
    if raw_source:
        try:
            raw_vanilla_keys = {
                item["normalized"] for item in parse_sitemap_anchors(raw_source, sitemap_url)
            }
        except ValueError:
            raw_vanilla_keys = set()

    existing = []
    existing_covered_normalized = set()
    
    for item in scoped_anchors:
        key = item["normalized"]
        uploaded = uploaded_map.get(key)
        status = uploaded["status"] if uploaded else ""
        
        final_url = ""
        if uploaded:
            status_group = _status_group(status)
            if status_group == "3xx" and uploaded.get("final_redirect_url"):
                final_url = uploaded["final_redirect_url"]
            else:
                final_url = item["url"]
        else:
            final_url = item["url"]

        if raw_vanilla_keys is None:
            in_source, source_type = "", ""
        elif key in raw_vanilla_keys:
            in_source, source_type = "Yes", "Vanilla HTML"
        else:
            in_source, source_type = "No", "JavaScript/Rendered DOM"
            
        existing.append(
            {
                "url": item["url"],
                "anchor": item["anchor"],
                "status": status,
                "in_sitemap": "Yes",
                "in_source": in_source,
                "source_type": source_type,
                "note": "",
                "final_url": final_url,
            }
        )
        existing_covered_normalized.add(key)
        if final_url:
            existing_covered_normalized.add(normalize_compare_url(final_url))

    missing_candidates = []
    for record in scoped_uploaded:
        if record["normalized"] in existing_covered_normalized:
            continue
        if _status_group(record["status"]) != "200":
            continue
        missing_candidates.append(record)

    missing = []
    for i, record in enumerate(missing_candidates, 1):
        missing.append(
            {
                "url": record["url"],
                "status": record["status"],
                "anchor": "",
            }
        )
        if progress:
            progress(i, len(missing_candidates))
    return existing, missing


def _apply_table_style(ws, widths: list[float], alignments: list[str]) -> None:
    blue_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = blue_fill
        cell.font = Font(name="Aptos Narrow", size=11, bold=True, color="000000")
        cell.border = border
        cell.alignment = Alignment(
            horizontal="left" if cell.column == 1 else "center",
            vertical="center",
            wrap_text=True,
        )
    for row in ws.iter_rows(min_row=2):
        for index, cell in enumerate(row):
            cell.font = Font(name="Aptos Narrow", size=11, color="000000")
            cell.border = border
            cell.alignment = Alignment(
                horizontal=alignments[index], vertical="center", wrap_text=False
            )
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + index)].width = width
    ws.auto_filter.ref = f"A1:{chr(64 + ws.max_column)}{max(1, ws.max_row)}"


def _build_audit_workbook(
    sitemap_url: str,
    existing: list[dict[str, object]],
    missing: list[dict[str, object]],
) -> bytes:
    template_path = Path(__file__).with_name("html_sitemap_template.xlsx")
    if not template_path.exists():
        raise ValueError("HTML sitemap Excel template is missing from the deployed app.")
    wb = load_workbook(template_path)
    summary = wb["Summary"]
    existing_ws = wb["Existing URLs in HTML Sitemap"]
    missing_ws = wb["URLs Not in HTML Sitemap"]

    summary["A1"] = f"Sitemap URL: {sitemap_url}"
    summary["B3"] = sum(_status_group(row["status"]) == "200" for row in existing)
    summary["C3"] = sum(_status_group(row["status"]) == "3xx" for row in existing)
    summary["D3"] = len(existing)
    summary["B4"] = len(missing)
    summary["C4"] = 0
    summary["D4"] = len(missing)

    def clear_values(ws, start_row: int) -> None:
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.value = None

    def ensure_styled_row(ws, target_row: int, template_row: int, columns: int) -> None:
        if target_row <= ws.max_row:
            return
        for column in range(1, columns + 1):
            source = ws.cell(template_row, column)
            target = ws.cell(target_row, column)
            target._style = copy(source._style)
            if source.has_style:
                target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.protection = copy(source.protection)
        ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height

    clear_values(existing_ws, 2)
    for index, row in enumerate(existing, 2):
        ensure_styled_row(existing_ws, index, 2, 8)
        values = [
            row["url"], row["anchor"], row["status"], row["in_sitemap"], row["in_source"],
            row["source_type"], "", row["final_url"],
        ]
        for column, value in enumerate(values, 1):
            existing_ws.cell(index, column).value = value
    existing_ws.auto_filter.ref = f"A1:H{max(1, len(existing) + 1)}"

    clear_values(missing_ws, 2)
    for index, row in enumerate(missing, 2):
        ensure_styled_row(missing_ws, index, 2, 3)
        missing_ws.cell(index, 1).value = row["url"]
        missing_ws.cell(index, 2).value = row["status"]
        missing_ws.cell(index, 3).value = row["anchor"]

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_html_sitemap_audit(
    url_status_data: bytes,
    url_status_filename: str,
    sitemap_source: str,
    sitemap_url: str,
    scope_mode: str,
    scope_pattern: str = "",
    queries: list[str] | None = None,
    api_key: str | None = None,
    progress=None,
) -> tuple[bytes, list[dict[str, object]], list[dict[str, object]], dict[str, int]]:
    uploaded = load_url_status_records(url_status_data, url_status_filename)
    anchors = parse_sitemap_anchors(sitemap_source, sitemap_url)
    existing, missing = _audit_rows(
        uploaded, anchors, sitemap_source, sitemap_url, scope_mode, scope_pattern,
        queries=queries, api_key=api_key, progress=progress,
    )
    output = _build_audit_workbook(sitemap_url, existing, missing)
    stats = {
        "uploaded_urls": sum(_in_scope(str(r["url"]), scope_mode, scope_pattern) for r in uploaded),
        "sitemap_links": len(existing),
        "missing_urls": len(missing),
        "sitemap_only": sum(str(r["status"]) == "" for r in existing),
    }
    return output, existing, missing, stats


def create_html_sitemap_audit_from_links(
    url_status_data: bytes,
    url_status_filename: str,
    rendered_links: list[dict[str, str]],
    sitemap_url: str,
    scope_mode: str,
    scope_pattern: str = "",
    raw_source: str | None = None,
    queries: list[str] | None = None,
    api_key: str | None = None,
    progress=None,
) -> tuple[bytes, list[dict[str, object]], list[dict[str, object]], dict[str, int]]:
    uploaded = load_url_status_records(url_status_data, url_status_filename)
    sitemap_host = (urlsplit(sitemap_url).hostname or "").casefold()
    anchors = []
    for item in rendered_links:
        url_value = str(item.get("url") or "").strip()
        if (urlsplit(url_value).hostname or "").casefold() != sitemap_host:
            continue
        anchors.append(
            {
                "url": url_value,
                "anchor": normalize_text(item.get("anchor", "")),
                "normalized": normalize_compare_url(url_value),
            }
        )
    if not anchors:
        raise ValueError("No same-domain rendered sitemap links are available for the audit.")
    existing, missing = _audit_rows(
        uploaded, anchors, raw_source, sitemap_url, scope_mode, scope_pattern,
        queries=queries, api_key=api_key, progress=progress,
    )
    output = _build_audit_workbook(sitemap_url, existing, missing)
    stats = {
        "uploaded_urls": sum(_in_scope(str(r["url"]), scope_mode, scope_pattern) for r in uploaded),
        "sitemap_links": len(existing),
        "missing_urls": len(missing),
        "sitemap_only": sum(str(r["status"]) == "" for r in existing),
    }
    return output, existing, missing, stats


def sitemap_output_filename(original_name: str) -> str:
    return f"{Path(original_name).stem}-HTML-Sitemap-Audit.xlsx"
