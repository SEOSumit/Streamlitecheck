import io
import math
import hashlib
import re
import markdownify
import time
import subprocess
import streamlit as st
import openpyxl
import concurrent.futures
import queue
from playwright.sync_api import sync_playwright, Error as PlaywrightError

def ensure_playwright_installed():
    """
    Attempts to run playwright browsers. If it fails due to missing executable,
    it runs `playwright install chromium`.
    """
    try:
        with sync_playwright() as p:
            # Just try to launch the browser quickly to see if it's installed
            browser = p.chromium.launch(headless=True)
            browser.close()
    except PlaywrightError as e:
        if "Executable doesn't exist" in str(e) or "playwright install" in str(e):
            st.info("First run detected: Installing Chromium for Playwright. This may take a minute...")
            subprocess.run(["playwright", "install", "chromium"], check=True)
            st.success("Chromium installed successfully!")
        else:
            raise e

def _process_chunk(chunk, progress_queue):
    """
    Processes a chunk of URLs in a single Playwright browser context.
    Yields progress back to the main thread via progress_queue.
    """
    results = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        # Add a realistic User-Agent to avoid getting blocked by Cloudflare/Akamai
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36"
        )

        for url in chunk:
            page = None
            extracted = False
            last_error_msg = ""
            
            for attempt in range(2): # Try up to 2 times
                try:
                    page = context.new_page()
                    # Increase timeout to 30s to help with slow JS pages
                    page.goto(url, wait_until='domcontentloaded', timeout=30000)
                    page.wait_for_timeout(1000) # 1s fixed wait instead of 2s
                    
                    # --- NEW LOGIC: Click Read More and FAQs, capturing dynamic Q&A ---
                    faq_extracted = []
                    try:
                        # 1. Click "Read More" buttons to reveal BOPC
                        read_more_btns = page.locator('button, [role="button"], .readMore, span.readTab')
                        count = read_more_btns.count()
                        for i in range(count):
                            btn = read_more_btns.nth(i)
                            try:
                                if btn.is_visible():
                                    text = btn.inner_text().strip().lower()
                                    if "read more" in text or "show more" in text:
                                        btn.click(timeout=500)
                                        page.wait_for_timeout(100)
                            except Exception:
                                pass

                        # 2. Click FAQs and capture their content immediately
                        accordion_btns = page.locator('.accordionContainer button, .faq button, .q-a button, [class*="faq" i] button, details summary')
                        count = accordion_btns.count()
                        for i in range(count):
                            btn = accordion_btns.nth(i)
                            try:
                                if btn.is_visible():
                                    text = btn.inner_text().strip()
                                    is_faq = "?" in text or btn.evaluate("el => el.closest('.accordionContainer, .faq, .q-a') !== null")
                                    if is_faq and len(text) > 5:
                                        btn.scroll_into_view_if_needed(timeout=500)
                                        btn.click(timeout=500)
                                        page.wait_for_timeout(200) # Wait for animation/render
                                    
                                        # Extract answer
                                        answer_text = ""
                                        controls = btn.get_attribute("aria-controls")
                                        if controls:
                                            region = page.locator(f'#{controls}')
                                            if region.count() > 0:
                                                answer_text = region.inner_text().strip()
                                    
                                        if not answer_text:
                                            answer_text = btn.evaluate("""el => {
                                                let next = el.nextElementSibling;
                                                if (!next && el.parentElement) {
                                                    next = el.parentElement.nextElementSibling;
                                                }
                                                // Handle specific nested structure in Lenovo accordions
                                                if (!next && el.parentElement && el.parentElement.parentElement) {
                                                    next = el.parentElement.parentElement.nextElementSibling;
                                                }
                                                return next ? next.innerText : '';
                                            }""")
                                    
                                        if answer_text and len(answer_text.strip()) > 5:
                                            faq_extracted.append(f"**Q: {text}**\\n\\nA: {answer_text.strip()}")
                            except Exception:
                                pass
                    except Exception:
                        pass
                
                    # Extract text content including hidden elements (like closed accordions)
                    # We remove scripts, styles, etc., to avoid polluting the text
                    # We also remove header, footer, nav, and aside to focus on the main body
                    extract_script = """
                    () => {
                        const clone = document.body.cloneNode(true);
                        const elementsToRemove = clone.querySelectorAll(
                            "script, style, noscript, iframe, link, meta, svg, " +
                            "header, footer, nav, aside, " +
                            "[role='navigation'], [role='banner'], [role='contentinfo'], " +
                            "#header, #footer"
                        );
                        for (let el of elementsToRemove) {
                            el.remove();
                        }
                        return clone.innerHTML;
                    }
                    """
                    raw_html = page.evaluate(extract_script)
                
                    if raw_html:
                        # Convert HTML to clean Markdown preserving headings and paragraphs
                        md_text = markdownify.markdownify(raw_html, heading_style="ATX", strip=['img', 'svg'])
                    
                        # Clean up excessive newlines
                        cleaned_text = re.sub(r'\\n{3,}', '\\n\\n', md_text).strip()
                    
                        # Append dynamically captured FAQs at the bottom if any were found
                        if faq_extracted:
                            # We use \n\n for proper markdown spacing
                            faq_section = "\\n\\n### Extracted FAQ Pairs:\\n\\n" + "\\n\\n".join(faq_extracted)
                            cleaned_text += faq_section
                        
                        if cleaned_text.strip():
                            results.append({"url": url, "text": cleaned_text, "error": None})
                            extracted = True
                            break # Success! Break out of the retry loop
                        else:
                            last_error_msg = "No text found in DOM"
                        
                except Exception as e:
                    # We specifically handle timeout errors to make the log clearer
                    error_str = str(e)
                    if "Timeout" in error_str:
                        last_error_msg = "Timeout: Page took too long to load"
                    elif "net::ERR_" in error_str:
                        last_error_msg = f"Network Error: {error_str.split('net::')[1].split(' at ')[0]}"
                    else:
                        last_error_msg = f"Error: {error_str.splitlines()[0]}"
                finally:
                    if page:
                        try:
                            page.close()
                        except Exception:
                            pass
                
                # If we get here, it means we failed. Wait a bit before retrying.
                if not extracted and attempt == 0:
                    time.sleep(2)

            if not extracted:
                results.append({"url": url, "text": None, "error": last_error_msg})
            
            # Delay between different URLs to avoid triggering rate limits
            time.sleep(0.5)
            progress_queue.put(1)
            
        browser.close()
    return results

def process_body_text_extractor(urls, progress_callback, max_workers):
    """
    Processes the list of URLs, extracts body text,
    and returns (markdown_text, stats).
    """
    total_urls = len(urls)
    if total_urls == 0:
        raise ValueError("No valid URLs provided to process.")

    ensure_playwright_installed()

    # Distribute URLs evenly across the available workers so concurrency is fully utilized
    # Ensure chunk size is at least 1, and no more than 50 (to prevent memory leaks)
    chunk_size = min(50, max(1, math.ceil(total_urls / max_workers)))
    chunks = [urls[i:i + chunk_size] for i in range(0, total_urls, chunk_size)]
    
    all_results = []
    progress_queue = queue.Queue()
    done_count = 0

    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(_process_chunk, chunk, progress_queue) for chunk in chunks]
        
        while done_count < total_urls:
            try:
                # Wait for a progress tick from any thread
                progress_queue.get(timeout=1.0)
                done_count += 1
                progress_callback(done_count, total_urls)
            except queue.Empty:
                # If queue is empty for a while, check if all futures are done (e.g. they crashed)
                if all(f.done() for f in futures):
                    break
        
        for f in futures:
            try:
                all_results.extend(f.result())
            except Exception:
                pass

    markdown_lines = []
    errors_list = []
    success_count = 0
    error_count = 0
    
    for res in all_results:
        url = res["url"]
        
        if res["text"]:
            markdown_lines.append(f"## {url}\n\n{res['text']}\n\n---")
            success_count += 1
        else:
            markdown_lines.append(f"## {url}\n\nERROR: {res['error']}\n\n---")
            errors_list.append({"URL": url, "Error": res["error"]})
            error_count += 1

    markdown_output = "\n".join(markdown_lines)
    
    stats = {
        "total": total_urls,
        "success": success_count,
        "error": error_count,
        "errors_list": errors_list
    }
    
    return markdown_output, stats

def run_extraction(urls, workers):
    if not urls:
        st.warning("No valid URLs found.")
        return
        
    progress_bar = st.progress(0, text="Initializing Playwright...")
    
    def update_progress(done: int, total: int) -> None:
        progress_bar.progress(done / total, text=f"Processing {done}/{total} URLs...")

    with st.spinner("Extracting body text. This may take a while depending on the number of URLs..."):
        try:
            output_md, stats = process_body_text_extractor(urls, update_progress, max_workers=workers)
        except Exception as exc:
            progress_bar.empty()
            st.error(str(exc))
        else:
            progress_bar.progress(1.0, text="Extraction complete")
            st.session_state["body_text_output_md"] = output_md
            st.session_state["body_text_stats"] = stats

def bulk_body_text_extractor_tool() -> None:
    st.title("Bulk Body Text Extractor")
    st.caption("Extract rendered body text from a list of URLs directly to an AI-friendly Markdown (.md) file.")
    
    mode = st.radio("Input Mode", ["Paste URL List", "Bulk Excel Upload"], horizontal=True)
    workers = st.slider("Max Concurrent Browsers", 1, 10, 5, help="Higher is faster but uses more memory. Reduce if the app crashes.")
    
    # Reset state if mode changes or new files uploaded
    def clear_state():
        for key in ("body_text_output_md", "body_text_stats"):
            st.session_state.pop(key, None)

    if mode == "Paste URL List":
        urls_text = st.text_area("Paste URLs (one per line)", on_change=clear_state)
        if st.button("Extract Body Text", type="primary", use_container_width=True):
            clear_state()
            urls = [u.strip() for u in urls_text.splitlines() if u.strip().startswith("http")]
            run_extraction(urls, workers)
            
    elif mode == "Bulk Excel Upload":
        with st.expander("Required sheet format"):
            st.markdown("Simply ensure that your **URLs are in the first column (Column A)** of the first sheet. Column headings or sheet names do not matter.")
            
        uploaded = st.file_uploader("Upload Excel file", type=["xlsx"], key="body_text_file", on_change=clear_state)
        
        if uploaded:
            st.markdown(
                '''
                <style>
                [data-testid="stFileUploadDropzone"] { display: none; }
                </style>
                ''',
                unsafe_allow_html=True,
            )
            if st.button("Extract Body Text", type="primary", use_container_width=True):
                clear_state()
                wb = openpyxl.load_workbook(io.BytesIO(uploaded.getvalue()))
                sheet = wb.worksheets[0]
                urls = []
                for row_idx in range(1, sheet.max_row + 1):
                    val = str(sheet.cell(row=row_idx, column=1).value).strip()
                    if val.startswith("http"):
                        urls.append(val)
                run_extraction(urls, workers)

    # Display results and download buttons
    if st.session_state.get("body_text_stats"):
        stats = st.session_state["body_text_stats"]
        col1, col2, col3 = st.columns(3)
        col1.metric("Total URLs Processed", stats["total"])
        col2.metric("Successfully Extracted", stats["success"])
        col3.metric("Errors", stats["error"])
        
        st.success("Extraction finished! Download your Markdown file below.")
        
        st.download_button(
            "Download Markdown (.md)",
            st.session_state["body_text_output_md"],
            "Extracted_Body_Text.md",
            "text/markdown",
            use_container_width=True,
            type="primary"
        )
        
        if stats["error"] > 0:
            st.divider()
            st.subheader(f"⚠️ {stats['error']} Errors Found")
            st.caption("The following URLs could not be extracted. This is usually caused by network blocks, timeouts, or the page returning 404/Access Denied.")
            
            # Display errors in a table
            st.dataframe(stats["errors_list"], use_container_width=True)
            
            # Provide error log download
            error_log_text = "\\n".join([f"{err['URL']} -> {err['Error']}" for err in stats["errors_list"]])
            st.download_button(
                "Download Error Log (.txt)",
                data=error_log_text,
                file_name="Extraction_Errors.txt",
                mime="text/plain"
            )
